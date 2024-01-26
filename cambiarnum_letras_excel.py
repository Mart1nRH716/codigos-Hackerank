from openpyxl import load_workbook

def number_to_words(number):
    # Diccionarios para convertir números a palabras
    ones = ['','uno', 'dos', 'tres', 'cuatro', 'cinco', 'seis', 'siete', 'ocho', 'nueve']
    teens = ['diez', 'once', 'doce', 'trece', 'catorce', 'quince', 'dieciséis', 'diecisiete', 'dieciocho', 'diecinueve']
    tens = ['','diez', 'veinte', 'treinta', 'cuarenta', 'cincuenta', 'sesenta', 'setenta', 'ochenta', 'noventa']
    hundreds = ['','ciento', 'doscientos', 'trescientos', 'cuatrocientos', 'quinientos', 'seiscientos', 'setecientos', 'ochocientos', 'novecientos']

    if number == 0:
        return 'cero'

    if number < 10:
        return ones[number]

    if number < 20:
        return teens[number - 10]

    if number < 100:
        return tens[number // 10] + ('' if number % 10 == 0 else ' y ' + ones[number % 10])

    if number < 1000:
        return hundreds[number // 100] + ('' if number % 100 == 0 else ' ' + number_to_words(number % 100))

    if number < 1000000:
        return number_to_words(number // 1000) + ' mil ' + number_to_words(number % 1000)

    if number < 1000000000:
        return number_to_words(number // 1000000) + ' millones ' + number_to_words(number % 1000000)

def main():
    # Carga el archivo Excel
    wb = load_workbook('tu_archivo.xlsx')
    # Escoge la primera hoja de cálculo
    sheet = wb.active

    # Itera sobre las celdas y convierte los números a palabras
    for row in sheet.iter_rows():
        for cell in row:
            if isinstance(cell.value, (int, float)):
                # Convierte el valor de la celda a palabras
                words = number_to_words(int(cell.value))
                # Imprime el valor convertido
                print(f"{cell.coordinate}: {words}")

if __name__ == "__main__":
    main()
